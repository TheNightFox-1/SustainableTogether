#!/usr/bin/env python3
"""validate_registry.py — machine V&V of the FBMC⇄CLD Concept Registry (invariants I1–I7).

Runs directly on the .xlsx (no RDF needed). Per Semantic-Integration-Playbook.md §A4
and TASK-BRIEF §P1. Importable API: `validate(path) -> report dict`. CLI:

    python3 validate_registry.py <registry.xlsx> [--json out.json]

Exit code 0 = all invariants pass (warnings allowed), 1 = at least one ERROR.

Severity policy (documented decisions):
- I2: an FBMC block from Lists with no concept at all -> ERROR, except "Context:*"
  blocks -> WARN (context is modelled selectively).
- I4: compound "and" in a CanonicalName -> WARN, suppressible by putting "I4-ok"
  in the Notes column (deliberate composite concept). Direction words -> ERROR.
- Instance rows (Notes: "instantiates GEN-xxx") are naming projections of a
  generic concept; they inherit link participation from that concept (I1, I6).
- I8 (FBMC->SD alignment, drawio page 5): SD well-formedness. Flow->Stock links
  must be unit-consistent (unit(flow) == unit(stock)+"/time"); Stocks fed by
  non-Flow variables need an intermediating Rate in SD. WARN-level until the
  unit policy is gated; escalate to ERROR once units are settled.
"""
import sys, re, json, datetime
from openpyxl import load_workbook

DIRECTION_WORDS = {"gain", "reduction", "pressure", "increase", "decrease",
                   "loss", "improvement", "decline", "shortfall", "surplus"}

REQUIRED_SHEETS = ["Concepts", "CausalLinks", "Loops", "ChangeLog", "Lists"]


def _rows(ws):
    """Yield dicts keyed by header row, skipping fully empty rows."""
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2):
        vals = [c.value for c in row]
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        yield {h: v for h, v in zip(headers, vals) if h}, row


def expand_id_tokens(text):
    """Expand ChangeLog ID references like 'GEN-001..GEN-021, L01..L34, R1..R4'."""
    ids = set()
    if not text:
        return ids
    for tok in re.split(r"[,;\s]+", str(text)):
        tok = tok.strip()
        if not tok:
            continue
        m = re.match(r"^([A-Za-z]+-?)(\d+)\.\.([A-Za-z]+-?)?(\d+)$", tok)
        if m:
            pre, a, pre2, b = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
            if pre2 is not None and pre2 != pre:
                continue  # malformed cross-prefix range (e.g. GEN-001..L34): expand nothing
            width = len(m.group(2))
            for i in range(a, b + 1):
                ids.add(f"{pre}{str(i).zfill(width)}")
        elif re.match(r"^[A-Za-z]+-?\d+$", tok):
            ids.add(tok)
    return ids


def validate(path):
    findings = []  # list of dicts: invariant, severity, message

    def err(inv, msg):  findings.append({"invariant": inv, "severity": "ERROR", "message": msg})
    def warn(inv, msg): findings.append({"invariant": inv, "severity": "WARN", "message": msg})

    wb_f = load_workbook(path, data_only=False)   # formulas visible
    # ---- S0 structural lint -------------------------------------------------
    for s in REQUIRED_SHEETS:
        if s not in wb_f.sheetnames:
            err("S0", f"required sheet missing: {s}")
    if any(f["severity"] == "ERROR" for f in findings):
        return _report(path, findings)

    concepts, links, loops = {}, {}, {}
    for d, _ in _rows(wb_f["Concepts"]):
        cid = d.get("ID")
        if not cid:
            continue
        if cid in concepts:
            err("S0", f"duplicate Concept ID {cid}")
        concepts[cid] = d
    for d, row in _rows(wb_f["CausalLinks"]):
        lid = d.get("LinkID")
        if not lid:
            continue
        if lid in links:
            err("S0", f"duplicate LinkID {lid}")
        links[lid] = d
        for col in ("FromName (auto)", "ToName (auto)"):
            v = d.get(col)
            if not (isinstance(v, str) and v.startswith("=")):
                err("S0", f"{lid}: '{col}' must be a lookup formula, found literal {v!r} "
                          "(asserted redundancy — Playbook §B rule 2)")
        if d.get("Polarity") not in ("+", "-", "−"):
            err("S0", f"{lid}: invalid Polarity {d.get('Polarity')!r}")
    for d, _ in _rows(wb_f["Loops"]):
        pid = d.get("LoopID")
        if not pid or not re.match(r"^[RB]\d+$", str(pid)):
            continue  # skip note rows
        if pid in loops:
            err("S0", f"duplicate LoopID {pid}")
        loops[pid] = d
        dt = d.get("DerivedType (formula)")
        if not (isinstance(dt, str) and dt.startswith("=") and "ISODD" in dt.upper()):
            err("S0", f"{pid}: DerivedType must be the ISODD formula, found {dt!r} "
                      "(loop type is derived, never asserted — invariant I3)")

    in_boundary = {cid for cid, d in concepts.items()
                   if str(d.get("InCLDBoundary", "")).strip().upper() == "Y"}

    # ---- I1 full trace / no orphans ----------------------------------------
    # Instance rows (Notes: "instantiates GEN-xxx") are naming projections of a
    # generic concept: the causal structure is defined ONCE over the generic IDs,
    # so an instance inherits link participation from the concept it instantiates.
    def instantiated(cid):
        m = re.search(r"instantiates\s+([A-Za-z]+-\d+)", str(concepts[cid].get("Notes") or ""))
        return m.group(1) if m else None

    linked = set()
    for lid, d in links.items():
        for col in ("FromID", "ToID"):
            ref = d.get(col)
            if ref not in concepts:
                err("I1", f"{lid}: {col} '{ref}' does not exist in Concepts")
            elif ref not in in_boundary:
                err("I1", f"{lid}: {col} '{ref}' is out of CLD boundary but used in a link")
            linked.add(ref)
    for cid in sorted(in_boundary):
        base = instantiated(cid)
        if base and base not in concepts:
            err("I1", f"{cid}: instantiates unknown concept '{base}'")
        if cid not in linked and (base or cid) not in linked:
            err("I1", f"{cid} is in CLD boundary but participates in no CausalLink (orphan variable)")
    for cid in sorted(in_boundary):
        if not str(concepts[cid].get("CLDVariableName") or "").strip():
            err("I1", f"{cid} is in boundary but has no CLDVariableName")

    # ---- I2 coverage per FBMC block -----------------------------------------
    lists_ws = wb_f["Lists"]
    blocks = [c.value for c in lists_ws["A"][1:] if c.value]
    by_block = {}
    for cid, d in concepts.items():
        by_block.setdefault(d.get("FBMC_Block"), []).append(cid)
    for b in blocks:
        if b not in by_block:
            (warn if str(b).startswith("Context") else err)(
                "I2", f"FBMC block '{b}' has no concept and no documented exclusion")
    for cid, d in concepts.items():
        if cid not in in_boundary and not str(d.get("ExclusionRationale") or "").strip():
            err("I2", f"{cid} is excluded from CLD but has no ExclusionRationale")

    # ---- I3 loop polarity derived & consistent ------------------------------
    for pid, d in loops.items():
        seq = [t.strip() for t in str(d.get("Ordered LinkIDs") or "").split(",") if t.strip()]
        chain = []
        for lid in seq:
            if lid not in links:
                err("I3", f"{pid}: unknown LinkID {lid}")
            else:
                chain.append(links[lid])
        if chain and len(chain) == len(seq):
            for i, lk in enumerate(chain):
                nxt = chain[(i + 1) % len(chain)]
                if lk.get("ToID") != nxt.get("FromID"):
                    err("I3", f"{pid}: chain broken at {seq[i]}→{seq[(i+1) % len(seq)]} "
                              f"({lk.get('ToID')} ≠ {nxt.get('FromID')})")
            neg = sum(1 for lk in chain if str(lk.get("Polarity")).strip() in ("-", "−"))
            manual = d.get("NegLinkCount (manual)")
            if manual is not None and int(manual) != neg:
                err("I3", f"{pid}: NegLinkCount is {manual} but recomputed value is {neg}")
            derived = "B" if neg % 2 == 1 else "R"
            if not str(pid).startswith(derived):
                err("I3", f"{pid}: recomputed type is {derived} "
                          f"({neg} negative links) but LoopID says {str(pid)[0]} "
                          "(v2-'R7' defect class)")

    # ---- I4 naming lint ------------------------------------------------------
    for cid, d in concepts.items():
        name = str(d.get("CanonicalName") or "")
        notes = str(d.get("Notes") or "")
        toks = set(re.findall(r"[a-z]+", name.lower()))
        hits = toks & DIRECTION_WORDS
        if hits:
            err("I4", f"{cid} '{name}': direction word(s) {sorted(hits)} — names must be "
                      "direction-neutral")
        if re.search(r"\band\b", name.lower()) and "I4-ok" not in notes:
            warn("I4", f"{cid} '{name}': compound name ('and') — split concept or mark "
                       "Notes with 'I4-ok' if composite by design")

    # ---- I5 actual ≠ perceived ----------------------------------------------
    names = {str(d.get("CanonicalName") or "").lower(): cid for cid, d in concepts.items()}
    for nm, cid in names.items():
        if "perceived" in nm:
            base = nm.replace("perceived", "").replace("  ", " ").strip()
            if not any(base in other and other != nm for other in names):
                err("I5", f"{cid} '{nm}': perceived variable has no actual counterpart "
                          "(actual-vs-perceived merge defect class)")

    # ---- I6 goal indicators present ------------------------------------------
    goal_ids = [cid for cid, d in concepts.items()
                if str(d.get("FBMC_Block") or "").startswith("Goals")]
    if not goal_ids:
        err("I6", "no concept anchored to the Goals (Outcomes) block")
    for cid in goal_ids:
        d = concepts[cid]
        if str(d.get("SDType")) != "Output":
            err("I6", f"{cid}: Goals concept must have SDType=Output, found {d.get('SDType')}")
        if cid not in in_boundary:
            err("I6", f"{cid}: goal indicator must be inside the CLD boundary")
        if cid not in linked and (instantiated(cid) or cid) not in linked:
            err("I6", f"{cid}: goal indicator appears in no CausalLink")

    # ---- I8 SD well-formedness (unit consistency, rate-only stock changes) ---
    def _norm_unit(u):
        return re.sub(r"\s+", " ", str(u or "").strip().lower())

    TIME_SUFFIXES = ("/year", "/yr", "/month", "/week", "/day")
    for lid, d in links.items():
        f_id, t_id = d.get("FromID"), d.get("ToID")
        if f_id not in concepts or t_id not in concepts:
            continue
        f_c, t_c = concepts[f_id], concepts[t_id]
        if str(t_c.get("SDType")) == "Stock":
            if str(f_c.get("SDType")) == "Flow":
                fu, tu = _norm_unit(f_c.get("Unit")), _norm_unit(t_c.get("Unit"))
                base = fu
                for suf in TIME_SUFFIXES:
                    if fu.endswith(suf):
                        base = fu[: -len(suf)]
                        break
                else:
                    warn("I8", f"{lid}: flow {f_id} unit '{f_c.get('Unit')}' has no per-time "
                               "suffix — a Rate must be measured per unit time")
                    continue
                if base != tu:
                    warn("I8", f"{lid}: unit mismatch — flow {f_id} '{f_c.get('Unit')}' should be "
                               f"stock {t_id} unit '{t_c.get('Unit')}' per time")
            else:
                warn("I8", f"{lid}: stock {t_id} is changed directly by non-Flow variable "
                           f"{f_id} ({f_c.get('SDType')}) — SD needs an intermediating Rate")

    # ---- I7 every ID in ChangeLog ---------------------------------------------
    logged = set()
    for d, _ in _rows(wb_f["ChangeLog"]):
        logged |= expand_id_tokens(d.get("ConceptIDs / LinkIDs / LoopIDs"))
    all_ids = set(concepts) | set(links) | set(loops)
    for missing in sorted(all_ids - logged):
        err("I7", f"{missing} never appears in the ChangeLog (all change via change records)")

    return _report(path, findings)


def _report(path, findings):
    errors = [f for f in findings if f["severity"] == "ERROR"]
    warns = [f for f in findings if f["severity"] == "WARN"]
    return {
        "registry": str(path),
        "timestamp": datetime.datetime.now().isoformat(timespec="seconds"),
        "errors": len(errors),
        "warnings": len(warns),
        "pass": not errors,
        "findings": findings,
    }


def main(argv):
    if len(argv) < 2:
        print(__doc__)
        return 2
    report = validate(argv[1])
    if "--json" in argv:
        out = argv[argv.index("--json") + 1]
        with open(out, "w", encoding="utf-8") as fh:
            json.dump(report, fh, indent=2, ensure_ascii=False)
    by_inv = {}
    for f in report["findings"]:
        by_inv.setdefault(f["invariant"], []).append(f)
    print(f"Registry V&V report — {report['registry']}  ({report['timestamp']})")
    print("-" * 72)
    for inv in ["S0", "I1", "I2", "I3", "I4", "I5", "I6", "I7", "I8"]:
        fs = by_inv.get(inv, [])
        status = "PASS" if not any(f["severity"] == "ERROR" for f in fs) else "FAIL"
        nwarn = sum(1 for f in fs if f["severity"] == "WARN")
        wtag = f" ({nwarn} warning(s))" if nwarn else ""
        print(f"{inv:4} {status}{wtag}")
        for f in fs:
            print(f"     [{f['severity']}] {f['message']}")
    print("-" * 72)
    print(f"RESULT: {'PASS' if report['pass'] else 'FAIL'} "
          f"- {report['errors']} error(s), {report['warnings']} warning(s)")
    return 0 if report["pass"] else 1


if __name__ == "__main__":
    sys.exit(main(sys.argv))

#!/usr/bin/env python3
"""xlsx2rdf.py — generate the ABox (registry.ttl) from the Concept Registry (L3).

Per Playbook §C4. Modelling decisions:
- Each registry Concept row yields ONE :MeasurableAttribute (:GEN-001) and, if
  in the CLD boundary, ONE :SystemVariable (:GEN-001-var) that :realizes it.
- FBMC blocks become :FBMCBlock instances (URI-safe local names).
- ChangeLog rows become :ChangeRecord (prov:Activity); every ID mentioned gets
  prov:wasGeneratedBy — which makes I7 SHACL-checkable.

    python3 xlsx2rdf.py <registry.xlsx> -o registry.ttl
"""
import re, argparse
from rdflib import Graph, Namespace, Literal, RDF, RDFS, XSD
from openpyxl import load_workbook

NS = Namespace("https://w3id.org/sustainabletogether/fbmc-cld#")
PROV = Namespace("http://www.w3.org/ns/prov#")
MINUS = "−"


def block_uri(block):
    return NS["Block_" + re.sub(r"[^A-Za-z0-9]+", "_", str(block)).strip("_")]


def expand_id_tokens(text):
    ids = set()
    for tok in re.split(r"[,;\s]+", str(text or "")):
        tok = tok.strip()
        m = re.match(r"^([A-Za-z]+-?)(\d+)\.\.([A-Za-z]+-?)?(\d+)$", tok)
        if m:
            if m.group(3) is not None and m.group(3) != m.group(1):
                continue  # malformed cross-prefix range: expand nothing
            width = len(m.group(2))
            for i in range(int(m.group(2)), int(m.group(4)) + 1):
                ids.add(f"{m.group(1)}{str(i).zfill(width)}")
        elif re.match(r"^[A-Za-z]+-?\d+$", tok):
            ids.add(tok)
    return ids


def rows(ws):
    heads = [c.value for c in ws[1]]
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(heads, r))
        if any(v is not None and str(v).strip() != "" for v in r):
            yield d


def convert(xlsx, out):
    wb = load_workbook(xlsx, data_only=True)
    g = Graph()
    g.bind("", NS)
    g.bind("prov", PROV)

    lit = lambda v, dt=None: Literal(str(v).strip(), datatype=dt) if v is not None and str(v).strip() else None

    def add(s, p, v, dt=None):
        l = lit(v, dt)
        if l is not None:
            g.add((s, p, l))

    # Blocks (from Lists col A)
    for c in wb["Lists"]["A"][1:]:
        if c.value:
            b = block_uri(c.value)
            g.add((b, RDF.type, NS.FBMCBlock))
            g.add((b, RDFS.label, Literal(str(c.value))))

    var_of = {}
    for d in rows(wb["Concepts"]):
        cid = d.get("ID")
        if not cid:
            continue
        c = NS[cid]
        g.add((c, RDF.type, NS.MeasurableAttribute))
        add(c, NS.canonicalName, d.get("CanonicalName"))
        add(c, NS.definition, d.get("Definition"))
        if d.get("FBMC_Block"):
            g.add((c, NS.isAttributeOf, block_uri(d["FBMC_Block"])))
        add(c, NS.unit, d.get("Unit"))
        add(c, NS.dimension, d.get("Dimension"))
        add(c, NS.sdType, d.get("SDType"))
        add(c, NS.exclusionRationale, d.get("ExclusionRationale"))
        add(c, NS.nameV1, d.get("Name_v1 (paper 1)"))
        add(c, NS.nameV2, d.get("Name_v2 (paper 2)"))
        add(c, NS.status, d.get("Status"))
        add(c, NS.version, d.get("Version"))
        inb = str(d.get("InCLDBoundary") or "").strip().upper() == "Y"
        g.add((c, NS.inCldBoundary, Literal(inb)))
        if inb:
            v = NS[cid + "-var"]
            var_of[cid] = v
            g.add((v, RDF.type, NS.SystemVariable))
            g.add((v, NS.realizes, c))
            add(v, NS.variableName, d.get("CLDVariableName") or d.get("CanonicalName"))

    for d in rows(wb["CausalLinks"]):
        lid = d.get("LinkID")
        if not lid:
            continue
        l = NS[lid]
        g.add((l, RDF.type, NS.CausalLink))
        for col, prop in (("FromID", NS["from"]), ("ToID", NS.to)):
            ref = str(d.get(col) or "")
            g.add((l, prop, var_of.get(ref, NS[ref + "-var"])))
        pol = str(d.get("Polarity") or "").strip()
        g.add((l, NS.polarity, Literal("-" if pol in ("-", MINUS) else "+")))
        g.add((l, NS.hasDelay, Literal(str(d.get("Delay") or "").strip().upper() == "Y")))
        add(l, NS.rationale, d.get("Rationale"))
        add(l, NS.status, d.get("Status"))

    for d in rows(wb["Loops"]):
        pid = str(d.get("LoopID") or "")
        if not re.match(r"^[RB]\d+$", pid):
            continue
        p = NS[pid]
        g.add((p, RDF.type, NS.FeedbackLoop))
        add(p, RDFS.label, d.get("LoopName"))
        add(p, NS.orderedLinkIds, d.get("Ordered LinkIDs"))
        add(p, NS.archetype, d.get("Archetype"))
        g.add((p, NS.loopType, Literal(pid[0])))  # derived upstream (I3 checks parity)
        for s in str(d.get("Ordered LinkIDs") or "").split(","):
            s = s.strip()
            if s:
                g.add((p, NS.hasLink, NS[s]))

    for i, d in enumerate(rows(wb["ChangeLog"]), 1):
        if not d.get("Date"):
            continue
        cr = NS[f"change-{i:03d}"]
        g.add((cr, RDF.type, NS.ChangeRecord))
        add(cr, PROV.startedAtTime, str(d.get("Date"))[:10], XSD.date)
        add(cr, NS.rationale, d.get("Description & rationale"))
        add(cr, PROV.wasAssociatedWith, d.get("Author"))
        for eid in expand_id_tokens(d.get("ConceptIDs / LinkIDs / LoopIDs")):
            g.add((NS[eid], PROV.wasGeneratedBy, cr))

    g.serialize(out, format="turtle")
    return g


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("registry")
    ap.add_argument("-o", "--out", default="registry.ttl")
    a = ap.parse_args()
    g = convert(a.registry, a.out)
    print(f"wrote {a.out}: {len(g)} triples")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""log_change.py — schema-enforced ChangeLog append (invariant I7 helper).

Agents and humans append change records through this script so entries can
never drift from the required schema. Per Playbook §B rule 3: the ChangeLog
entry lands BEFORE any artifact is republished.

    python3 log_change.py <registry.xlsx> \
        --type add|rename|split|merge|retire|link-change|loop-change \
        --ids "PV-001..PV-021, L35" \
        --desc "why" --author "name" \
        [--version 0.2] [--artifacts "CLD v3, registry.ttl"] [--date YYYY-MM-DD]

Exit 0 on success; validates type against the Lists sheet and refuses empty fields.
"""
import sys, argparse, datetime
from openpyxl import load_workbook


def append_change(path, change_type, ids, desc, author,
                  version=None, artifacts="", date=None):
    wb = load_workbook(path)
    for sheet in ("ChangeLog", "Lists"):
        if sheet not in wb.sheetnames:
            raise SystemExit(f"ERROR: sheet '{sheet}' missing in {path}")
    allowed = [c.value for c in wb["Lists"]["F"][1:] if c.value]
    if change_type not in allowed:
        raise SystemExit(f"ERROR: ChangeType '{change_type}' not in Lists: {allowed}")
    for field, val in [("ids", ids), ("desc", desc), ("author", author)]:
        if not str(val).strip():
            raise SystemExit(f"ERROR: --{field} must not be empty")

    ws = wb["ChangeLog"]
    if version is None:  # reuse last version if not given
        version = ws.cell(ws.max_row, 2).value or "0.1"
    row = [date or datetime.date.today().isoformat(), str(version), change_type,
           ids, desc, author, artifacts]
    ws.append(row)
    wb.save(path)
    print(f"ChangeLog: appended row {ws.max_row}: {row}")


def main():
    p = argparse.ArgumentParser()
    p.add_argument("registry")
    p.add_argument("--type", required=True)
    p.add_argument("--ids", required=True)
    p.add_argument("--desc", required=True)
    p.add_argument("--author", required=True)
    p.add_argument("--version", default=None)
    p.add_argument("--artifacts", default="")
    p.add_argument("--date", default=None)
    a = p.parse_args()
    append_change(a.registry, a.type, a.ids, a.desc, a.author,
                  a.version, a.artifacts, a.date)


if __name__ == "__main__":
    main()

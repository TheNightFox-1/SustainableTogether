#!/usr/bin/env python3
"""registry2cld.py — project the Concept Registry into a draw.io CLD page.

The CLD is a GENERATED artifact (Playbook §A5: regenerate, never hand-edit).
Layout strategy, in order of preference per node:
  1. exact CLDVariableName match against a template page (default: page 4 of
     FBMC-CLD-Alignment-Ontology.drawio);
  2. 'instantiates GEN-xxx' marker in the Notes column -> inherit that GEN
     concept's template position (used by PV- instance rows);
  3. fallback: circle layout for anything unmatched.
Loop badges (R/B ellipses) are placed at the centroid of their member nodes.

    python3 registry2cld.py <registry.xlsx> -o out.drawio
        [--template ../FBMC-CLD-Alignment-Ontology.drawio --template-page page-generic-cld]
        [--verify]

--verify re-parses the output and checks node/edge/polarity fidelity vs the
registry (round-trip V&V per TASK-BRIEF §3 station 6). Exit 1 on mismatch.
"""
import sys, os, re, argparse, datetime, html
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
DEF_TEMPLATE = os.path.join(HERE, "..", "FBMC-CLD-Alignment-Ontology.drawio")

NODE_STYLE = ("rounded=1;whiteSpace=wrap;html=1;fillColor=none;strokeColor=none;"
              "fontSize=12;fontStyle=1;fontColor={color};")
EDGE_STYLE = ("curved=1;html=1;endArrow=classic;endSize=8;strokeColor=#666666;"
              "fontSize=14;fontStyle=1;{extra}")
LOOP_STYLE = ("ellipse;fillColor=none;strokeColor={color};fontSize=13;"
              "fontStyle=1;fontColor={color};")
MINUS = "−"  # the − glyph used in the template


def read_registry(path, instance_prefix=None):
    """Instance rows (Notes: 'instantiates GEN-xxx') are naming projections,
    never separate nodes. Default: generic view (generic names). With
    instance_prefix (e.g. 'PV'): same structure, instance names substituted."""
    wb = load_workbook(path, data_only=True)
    concepts, links, loops = {}, [], []
    instances = {}
    heads = [c.value for c in wb["Concepts"][1]]
    for row in wb["Concepts"].iter_rows(min_row=2, values_only=True):
        d = dict(zip(heads, row))
        if not d.get("ID") or str(d.get("InCLDBoundary", "")).strip().upper() != "Y":
            continue
        m = re.search(r"instantiates\s+([A-Za-z]+-\d+)", str(d.get("Notes") or ""))
        if m:
            instances.setdefault(m.group(1), []).append(d)
            continue
        concepts[d["ID"]] = d
    if instance_prefix:
        for base, cands in instances.items():
            hit = next((c for c in cands if str(c["ID"]).startswith(instance_prefix)), None)
            if hit and base in concepts:
                g = dict(concepts[base])
                g["_generic_name"] = str(g.get("CLDVariableName") or g.get("CanonicalName"))
                g["CLDVariableName"] = hit.get("CLDVariableName") or hit.get("CanonicalName")
                g["CanonicalName"] = hit.get("CanonicalName") or g["CanonicalName"]
                g["FBMC_Element (instance)"] = hit.get("FBMC_Element (instance)")
                concepts[base] = g
    heads = [c.value for c in wb["CausalLinks"][1]]
    for row in wb["CausalLinks"].iter_rows(min_row=2, values_only=True):
        d = dict(zip(heads, row))
        if d.get("LinkID"):
            links.append(d)
    heads = [c.value for c in wb["Loops"][1]]
    for row in wb["Loops"].iter_rows(min_row=2, values_only=True):
        d = dict(zip(heads, row))
        if d.get("LoopID") and re.match(r"^[RB]\d+$", str(d["LoopID"])):
            loops.append(d)
    return concepts, links, loops


def read_template(path, page_id):
    """Return {node_label_first_line: (x, y, w, h, fontColor)} and loop positions."""
    pos, loop_pos = {}, {}
    if not os.path.exists(path):
        return pos, loop_pos
    tree = ET.parse(path)
    for d in tree.getroot().findall("diagram"):
        if d.get("id") != page_id and d.get("name") != page_id:
            continue
        for c in d.iter("mxCell"):
            geo = c.find("mxGeometry")
            if geo is None or c.get("edge") == "1":
                continue
            val = c.get("value") or ""
            style = c.get("style") or ""
            g = (float(geo.get("x", 0)), float(geo.get("y", 0)),
                 float(geo.get("width", 120)), float(geo.get("height", 40)))
            if style.startswith("ellipse"):
                loop_pos[val.strip()] = g
            elif val and "rounded=1" in style:
                label = re.split(r"[\[\n]", html.unescape(val))[0].strip()
                m = re.search(r"fontColor=(#[0-9a-fA-F]{6})", style)
                pos[label.lower()] = g + ((m.group(1) if m else "#000000"),)
    return pos, loop_pos


def build(registry, out, template, template_page, title=None, instance_prefix=None):
    concepts, links, loops = read_registry(registry, instance_prefix)
    tpos, tloops = read_template(template, template_page)

    # --- resolve positions ---------------------------------------------------
    place, unplaced = {}, []
    for cid, d in concepts.items():
        name = str(d.get("CLDVariableName") or d.get("CanonicalName")).strip()
        key = name.lower()
        hit = tpos.get(key)
        if not hit and d.get("_generic_name"):
            hit = tpos.get(str(d["_generic_name"]).strip().lower())  # instance view: generic layout
        if not hit:
            m = re.search(r"instantiates\s+([A-Z]+-\d+)", str(d.get("Notes") or ""))
            if m and m.group(1) in concepts:
                src = concepts[m.group(1)]
                hit = tpos.get(str(src.get("CLDVariableName") or "").strip().lower())
        if hit:
            place[cid] = hit
        else:
            unplaced.append(cid)
    if unplaced:  # circle fallback
        import math
        cx, cy, r = 800, 550, 430
        for i, cid in enumerate(unplaced):
            a = 2 * math.pi * i / len(unplaced)
            place[cid] = (cx + r * math.cos(a), cy + r * math.sin(a), 170, 40, "#000000")

    # --- XML -----------------------------------------------------------------
    mxfile = ET.Element("mxfile", host="app.diagrams.net")
    diagram = ET.SubElement(mxfile, "diagram", id="generated-cld",
                            name=title or "CLD (generated from registry)")
    model = ET.SubElement(diagram, "mxGraphModel", dx="1000", dy="800", grid="0",
                          gridSize="10", guides="1", tooltips="1", connect="1",
                          arrows="1", fold="1", page="1", pageScale="1",
                          pageWidth="1650", pageHeight="1100", math="0", shadow="0")
    root = ET.SubElement(model, "root")
    ET.SubElement(root, "mxCell", id="0")
    ET.SubElement(root, "mxCell", id="1", parent="0")

    stamp = datetime.date.today().isoformat()
    t = ET.SubElement(root, "mxCell", id="t1", parent="1", vertex="1",
                      style="text;html=1;fontSize=14;fontStyle=1;align=left;",
                      value=f"CLD generated from {os.path.basename(registry)} on {stamp} "
                            "— DO NOT HAND-EDIT (regenerate via registry2cld.py)")
    ET.SubElement(t, "mxGeometry", x="40", y="10", width="1500", height="30",
                  **{"as": "geometry"})

    for cid, d in concepts.items():
        x, y, w, h, color = place[cid]
        name = str(d.get("CLDVariableName") or d.get("CanonicalName"))
        block = str(d.get("FBMC_Block") or "").split(" (")[0]
        sdt = d.get("SDType") or ""
        cell = ET.SubElement(root, "mxCell", id=cid, parent="1", vertex="1",
                             style=NODE_STYLE.format(color=color),
                             value=f"{name}\n[{block} • {sdt}]")
        ET.SubElement(cell, "mxGeometry", x=str(x), y=str(y), width=str(w),
                      height=str(h), **{"as": "geometry"})

    for d in links:
        pol = str(d.get("Polarity")).strip()
        neg = pol in ("-", MINUS)
        extra = "fontColor=#cc0000;" if neg else ""
        if str(d.get("Delay") or "").strip().upper() == "Y":
            extra += "dashed=1;"
        cell = ET.SubElement(root, "mxCell", id=str(d["LinkID"]), parent="1",
                             edge="1", source=str(d["FromID"]), target=str(d["ToID"]),
                             style=EDGE_STYLE.format(extra=extra),
                             value=MINUS if neg else "+")
        ET.SubElement(cell, "mxGeometry", relative="1", **{"as": "geometry"})

    for d in loops:
        lid = str(d["LoopID"])
        seq = [s.strip() for s in str(d.get("Ordered LinkIDs") or "").split(",") if s.strip()]
        members = set()
        by_id = {str(l["LinkID"]): l for l in links}
        for s in seq:
            if s in by_id:
                members |= {by_id[s]["FromID"], by_id[s]["ToID"]}
        pts = [place[m] for m in members if m in place]
        if lid in tloops:
            x, y, w, h = tloops[lid]
        elif pts:
            x = sum(p[0] for p in pts) / len(pts) + 60
            y = sum(p[1] for p in pts) / len(pts)
            w, h = 44, 44
        else:
            continue
        color = "#a33333" if lid.startswith("B") else "#3b6ba5"
        cell = ET.SubElement(root, "mxCell", id=f"lp-{lid}", parent="1", vertex="1",
                             style=LOOP_STYLE.format(color=color),
                             value=f"{lid}\n{d.get('LoopName') or ''}")
        ET.SubElement(cell, "mxGeometry", x=str(x), y=str(y), width=str(w),
                      height=str(h), **{"as": "geometry"})

    ET.ElementTree(mxfile).write(out, encoding="utf-8", xml_declaration=True)
    return concepts, links, loops


def verify(out, concepts, links):
    """Round-trip: parse output, compare against registry. Returns list of problems."""
    problems = []
    tree = ET.parse(out)
    nodes = {c.get("id"): c for c in tree.iter("mxCell")
             if c.get("vertex") == "1" and c.get("id") in concepts}
    edges = {c.get("id"): c for c in tree.iter("mxCell") if c.get("edge") == "1"}
    if set(nodes) != set(concepts):
        problems.append(f"node set mismatch: missing {set(concepts)-set(nodes)}, "
                        f"extra {set(nodes)-set(concepts)}")
    reg_edges = {str(l["LinkID"]): l for l in links}
    if set(edges) != set(reg_edges):
        problems.append(f"edge set mismatch: missing {set(reg_edges)-set(edges)}, "
                        f"extra {set(edges)-set(reg_edges)}")
    for lid, e in edges.items():
        r = reg_edges.get(lid)
        if not r:
            continue
        if e.get("source") != str(r["FromID"]) or e.get("target") != str(r["ToID"]):
            problems.append(f"{lid}: endpoints {e.get('source')}->{e.get('target')} "
                            f"≠ registry {r['FromID']}->{r['ToID']}")
        want = MINUS if str(r["Polarity"]).strip() in ("-", MINUS) else "+"
        if (e.get("value") or "").strip() != want:
            problems.append(f"{lid}: polarity label {e.get('value')!r} ≠ {want!r}")
    return problems


def main():
    p = argparse.ArgumentParser()
    p.add_argument("registry")
    p.add_argument("-o", "--out", default="cld_generated.drawio")
    p.add_argument("--template", default=DEF_TEMPLATE)
    p.add_argument("--template-page", default="page-generic-cld")
    p.add_argument("--title", default=None)
    p.add_argument("--instance-prefix", default=None,
                   help="project the instance view, e.g. PV (names from PV- rows on generic structure)")
    p.add_argument("--verify", action="store_true")
    a = p.parse_args()
    concepts, links, loops = build(a.registry, a.out, a.template, a.template_page, a.title,
                                   a.instance_prefix)
    print(f"wrote {a.out}: {len(concepts)} variables, {len(links)} links, {len(loops)} loop badges")
    if a.verify:
        problems = verify(a.out, concepts, links)
        if problems:
            print("ROUND-TRIP FAIL:")
            for x in problems:
                print("  -", x)
            sys.exit(1)
        print("round-trip verification: PASS")


if __name__ == "__main__":
    main()

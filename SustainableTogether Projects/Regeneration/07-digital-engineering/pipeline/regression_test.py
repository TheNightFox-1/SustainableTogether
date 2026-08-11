#!/usr/bin/env python3
"""regression_test.py — proves validate_registry.py catches every known defect class.

Takes the clean registry, applies one seeded defect at a time (in a temp copy),
and asserts the validator FAILS with the expected invariant. Also asserts the
clean registry PASSES. Fixture xlsx files are written to fixtures/ for inspection.

    python3 regression_test.py [registry.xlsx]

Exit 0 = all regression cases behave as expected.
"""
import sys, os, shutil, re
from openpyxl import load_workbook
import validate_registry as vr

HERE = os.path.dirname(os.path.abspath(__file__))
FIXDIR = os.path.join(HERE, "fixtures")


def _col(ws, header):
    for i, c in enumerate(ws[1], 1):
        if c.value == header:
            return i
    raise KeyError(header)


def _find_row(ws, col_idx, value):
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, col_idx).value == value:
            return r
    raise KeyError(value)


# --- defect mutations (each mirrors an observed v1/v2 defect class) -----------

def d_orphan_variable(wb):
    """I1: add an in-boundary concept that no link touches."""
    ws = wb["Concepts"]
    r = ws.max_row + 1
    ws.cell(r, _col(ws, "ID"), "GEN-099")
    ws.cell(r, _col(ws, "CanonicalName"), "dangling test variable")
    ws.cell(r, _col(ws, "FBMC_Block"), "Actors (Who)")
    ws.cell(r, _col(ws, "SDType"), "Auxiliary")
    ws.cell(r, _col(ws, "InCLDBoundary"), "Y")
    ws.cell(r, _col(ws, "CLDVariableName"), "dangling test variable")


def d_dangling_link(wb):
    """I1: link referencing a nonexistent concept."""
    ws = wb["CausalLinks"]
    r = _find_row(ws, _col(ws, "LinkID"), "L01")
    ws.cell(r, _col(ws, "ToID"), "GEN-777")


def d_missing_exclusion(wb):
    """I2: excluded concept without rationale."""
    ws = wb["Concepts"]
    r = _find_row(ws, _col(ws, "ID"), "GEN-021")
    ws.cell(r, _col(ws, "InCLDBoundary"), "N")
    ws.cell(r, _col(ws, "ExclusionRationale"), None)
    # keep links referencing GEN-021 — that also violates I1, but I2 must fire


def d_wrong_loop_polarity(wb):
    """I3: the v2-'R7' defect — loop labelled R whose recomputed type is B."""
    ws = wb["Loops"]
    r = _find_row(ws, _col(ws, "LoopID"), "B1")
    ws.cell(r, _col(ws, "LoopID"), "R9")          # mislabel as reinforcing
    ws.cell(r, _col(ws, "NegLinkCount (manual)"), 0)  # and miscount


def d_asserted_loop_type(wb):
    """S0: DerivedType typed literally instead of formula."""
    ws = wb["Loops"]
    r = _find_row(ws, _col(ws, "LoopID"), "R1")
    ws.cell(r, _col(ws, "DerivedType (formula)"), "R")


def d_direction_word_name(wb):
    """I4: direction word in a canonical name."""
    ws = wb["Concepts"]
    r = _find_row(ws, _col(ws, "ID"), "GEN-002")
    ws.cell(r, _col(ws, "CanonicalName"), "need satisfaction increase")


def d_perceived_without_actual(wb):
    """I5: perceived variable with no actual counterpart."""
    ws = wb["Concepts"]
    r = _find_row(ws, _col(ws, "ID"), "GEN-017")
    for col in ("CanonicalName", "CLDVariableName"):
        ws.cell(r, _col(ws, col), "perceived ecosystem service capacity")


def d_goal_not_output(wb):
    """I6: Goals concept demoted to Auxiliary."""
    ws = wb["Concepts"]
    r = _find_row(ws, _col(ws, "ID"), "GEN-020")
    ws.cell(r, _col(ws, "SDType"), "Auxiliary")


def d_unlogged_id(wb):
    """I7: new link never mentioned in ChangeLog."""
    ws = wb["CausalLinks"]
    r = ws.max_row + 1
    ws.cell(r, _col(ws, "LinkID"), "L99")
    ws.cell(r, _col(ws, "FromID"), "GEN-001")
    ws.cell(r, _col(ws, "ToID"), "GEN-005")
    ws.cell(r, _col(ws, "FromName (auto)"), '=IFERROR(VLOOKUP(B99,Concepts!$A:$B,2,FALSE()),"")')
    ws.cell(r, _col(ws, "ToName (auto)"), '=IFERROR(VLOOKUP(C99,Concepts!$A:$B,2,FALSE()),"")')
    ws.cell(r, _col(ws, "Polarity"), "+")


def d_broken_chain(wb):
    """I3: loop whose links do not connect head-to-tail."""
    ws = wb["Loops"]
    r = _find_row(ws, _col(ws, "LoopID"), "R1")
    ws.cell(r, _col(ws, "Ordered LinkIDs"), "L09,L10,L06,L05,L01")  # L04 removed


def d_sd_unit_mismatch(wb):
    """I8: flow unit no longer stock-unit-per-time."""
    ws = wb["Concepts"]
    r = _find_row(ws, _col(ws, "ID"), "GEN-005")   # flow into stock GEN-003
    ws.cell(r, _col(ws, "Unit"), "EUR/year")       # stock is 'actors'


WARN_CASES = [
    # (name, mutation, invariant expected among WARN findings)
    ("sd_unit_mismatch", d_sd_unit_mismatch, "I8"),
]

CASES = [
    ("orphan_variable",         d_orphan_variable,        "I1"),
    ("dangling_link",           d_dangling_link,          "I1"),
    ("missing_exclusion",       d_missing_exclusion,      "I2"),
    ("wrong_loop_polarity",     d_wrong_loop_polarity,    "I3"),
    ("broken_chain",            d_broken_chain,           "I3"),
    ("asserted_loop_type",      d_asserted_loop_type,     "S0"),
    ("direction_word_name",     d_direction_word_name,    "I4"),
    ("perceived_without_actual", d_perceived_without_actual, "I5"),
    ("goal_not_output",         d_goal_not_output,        "I6"),
    ("unlogged_id",             d_unlogged_id,            "I7"),
]


def main(argv):
    registry = argv[1] if len(argv) > 1 else os.path.join(HERE, "..", "SustainaSun_Concept_Registry.xlsx")
    os.makedirs(FIXDIR, exist_ok=True)
    failures = []

    clean = vr.validate(registry)
    ok = clean["pass"]
    print(f"clean registry               : {'PASS' if ok else 'FAIL'} (expected PASS)")
    if not ok:
        failures.append("clean")

    for name, mutate, expected_inv in CASES:
        fix = os.path.join(FIXDIR, f"defect_{name}.xlsx")
        shutil.copy(registry, fix)
        wb = load_workbook(fix)
        mutate(wb)
        wb.save(fix)
        rep = vr.validate(fix)
        fired = {f["invariant"] for f in rep["findings"] if f["severity"] == "ERROR"}
        ok = (not rep["pass"]) and expected_inv in fired
        print(f"defect_{name:24}: {'CAUGHT' if ok else 'MISSED'} "
              f"(expected {expected_inv}, errors fired: {sorted(fired) or 'none'})")
        if not ok:
            failures.append(name)

    for name, mutate, expected_inv in WARN_CASES:
        fix = os.path.join(FIXDIR, f"defect_{name}.xlsx")
        shutil.copy(registry, fix)
        wb = load_workbook(fix)
        mutate(wb)
        wb.save(fix)
        rep = vr.validate(fix)
        fired = {f["invariant"] for f in rep["findings"] if f["severity"] == "WARN"}
        ok = expected_inv in fired
        print(f"defect_{name:24}: {'CAUGHT' if ok else 'MISSED'} "
              f"(expected WARN {expected_inv}, warns fired: {sorted(fired) or 'none'})")
        if not ok:
            failures.append(name)

    total = len(CASES) + len(WARN_CASES) + 1
    print("-" * 72)
    print(f"regression: {total - len(failures)}/{total} cases OK")
    return 1 if failures else 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
